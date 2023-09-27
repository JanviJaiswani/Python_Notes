from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.common.exceptions import NoSuchElementException
import time

# Create a WebDriver instance for Edge
driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()))

# Define the website URL
url = "https://demowebshop.tricentis.com/login"

# Load the Excel file
excel_file = "records_users.xlsx"
workbook = load_workbook(excel_file)
worksheet = workbook.active

for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
    id, password, Validity = row

    driver.get(url)

    # Find the username and password fields and fill them
    username_field = driver.find_element(By.ID, "Email")
    password_field = driver.find_element(By.ID, "Password")

    username_field.send_keys(id)
    password_field.send_keys(password)

    # Click the login button
    login_button = driver.find_element(By.CSS_SELECTOR, "input[value='Log in']")
    login_button.click()
    time.sleep(2)

    try:
        # Attempt to find the "Log out" link
        logout_link = driver.find_element(By.LINK_TEXT, "Log out")

        # If the link is found, the login was successful
        Validity = "Yes"

        # Log out the user
        logout_link.click()

    except NoSuchElementException:
        # If the "Log out" link is not found, the login was not successful
        Validity = "No"

    # Update the "Valid User" column in the Excel file for the current row
    worksheet.cell(row=row_number, column=3, value=Validity)

# Save the updated Excel file
workbook.save(excel_file)

# Close the WebDriver
driver.quit()
